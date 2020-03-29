import urllib,urllib2, os, openpyxl, sys
from bs4 import  BeautifulSoup
from mechanize import Browser
from datetime import datetime
from openpyxl import Workbook
import pandas as pd

start = datetime.now()

print "Downloading datebase.....\n\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book = openpyxl.load_workbook("NSE-BSE list.xlsx")
scripts = book.get_sheet_by_name("NSE")

br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker\NewerDB")
not_listed = []

for i in range(1,scripts.max_row+1):
    script_id = scripts['B'+str(i)].value
        
    try:
        response=br.open("https://in.finance.yahoo.com/quote/"+script_id+".NS/history?period1=1486146600&period2=1498415400&interval=1d&filter=history&frequency=1d")
        if br.title()=="Symbol lookup from Yahoo Finance":
            not_listed.append(script_id)
            continue
    except:
        not_listed.append(script_id)
        continue
    soup = BeautifulSoup(response.read(),"html.parser")
    print br.title()
    right_table = soup.find( "table", { "class" : "W(100%) M(0) BdB Bdc($lightGray)" } )

    months = { "Feb":'02', "Mar":'03',"Apr":'04',"May":'05',"Jun":'06'}
    dates = []
    open = []
    close = []
    adj = []
    vol =[]
    high = []
    low = [] 
    
    for rows in right_table.findAll("tr"):
        cells = rows.findAll("td")
        if len(cells)==7:
            date = str(cells[0].text)
            dd = date[:2]
            mm = months[date[3:6]]
            yy = date[7:]
            dates.append(yy+"-"+mm+"-"+dd)
            a = str(cells[1].text)
            a = a.replace(",","")
            a = float(a)
            open.append(a)
            a = str(cells[2].text)
            a = a.replace(",","")
            a = float(a)
            high.append(a)
            a = str(cells[3].text)
            a = a.replace(",","")
            a = float(a)
            low.append(a)
            a = str(cells[4].text)
            a = a.replace(",","")
            a = float(a)
            close.append(a)
            a = str(cells[5].text)
            a = a.replace(",","")
            a = float(a)
            adj.append(a)
            a = str(cells[6].text)
            a = a.replace(",","")
            if len(a)==1:
                a = 0
            else:
                a = float(a)
            vol.append(a)
        
    df = pd.DataFrame(dates,columns=['Date'])
    df['Open']=open
    df['High']=high
    df['Low']=low
    df['Close']=close
    df['Volume']=vol
    df['Adj Close']=adj
    
    writer = pd.ExcelWriter(script_id+".xlsx")
    #df = pandas.read_excel('your_xls_xlsx_filename', sheetname='Sheet 1')   
    os.chdir("C:\Users\dell\Documents\Stock Picker\DB")
    df1 = pd.read_excel(script_id+".xlsx", sheetname='Sheet1')
    dataframe = pd.concat([df,df1])
    dataframe.to_excel(writer,index=False)
    os.chdir("C:\Users\dell\Documents\Stock Picker\NewerDB")
    writer.save()
    print script_id+".xlsx"+" is ready!"
   
scripts = book.get_sheet_by_name("BSE")

br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker\NewDB")
not_listed = []

for i in range(1,scripts.max_row+1):
    script_id = scripts['B'+str(i)].value
        
    try:
        response=br.open("https://in.finance.yahoo.com/quote/"+script_id+".BO/history?period1=1486146600&period2=1498415400&interval=1d&filter=history&frequency=1d")
        if br.title()=="Symbol lookup from Yahoo Finance":
            not_listed.append(script_id)
            continue
    except:
        not_listed.append(script_id)
        continue
    soup = BeautifulSoup(response.read(),"html.parser")
    print br.title()
    right_table = soup.find( "table", { "class" : "W(100%) M(0) BdB Bdc($lightGray)" } )

    months = { "Feb":'02', "Mar":'03',"Apr":'04',"May":'05',"Jun":'06'}
    dates = []
    open = []
    close = []
    adj = []
    vol =[]
    high = []
    low = [] 
    
    for rows in right_table.findAll("tr"):
        cells = rows.findAll("td")
        if len(cells)==7:
            date = str(cells[0].text)
            dd = date[:2]
            mm = months[date[3:6]]
            yy = date[7:]
            dates.append(yy+"-"+mm+"-"+dd)
            a = str(cells[1].text)
            a = a.replace(",","")
            a = float(a)
            open.append(a)
            a = str(cells[2].text)
            a = a.replace(",","")
            a = float(a)
            high.append(a)
            a = str(cells[3].text)
            a = a.replace(",","")
            a = float(a)
            low.append(a)
            a = str(cells[4].text)
            a = a.replace(",","")
            a = float(a)
            close.append(a)
            a = str(cells[5].text)
            a = a.replace(",","")
            a = float(a)
            adj.append(a)
            a = str(cells[6].text)
            a = a.replace(",","")
            if len(a)==1:
                a = 0
            else:
                a = float(a)
            vol.append(a)
        
    df = pd.DataFrame(dates,columns=['Date'])
    df['Open']=open
    df['High']=high
    df['Low']=low
    df['Close']=close
    df['Volume']=vol
    df['Adj Close']=adj
    
    writer = pd.ExcelWriter(script_id+".xlsx")
    #df = pandas.read_excel('your_xls_xlsx_filename', sheetname='Sheet 1')   
    os.chdir("C:\Users\dell\Documents\Stock Picker\DB")
    df1 = pd.read_excel(script_id+".xlsx", sheetname='Sheet1')
    dataframe = pd.concat([df,df1])
    dataframe.to_excel(writer,index=False)
    os.chdir("C:\Users\dell\Documents\Stock Picker\NewDB")
    writer.save()
    print script_id+".xlsx"+" is ready!"
    
 
'''

# <----------------------52 week low high----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("52-week low.py")

# <----------------------Simple Moving Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Simple Moving Average.py")

# <----------------------Exponential Moving Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Exponential Moving Average.py")

# <----------------------Volume Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Volume Average.py")

'''






